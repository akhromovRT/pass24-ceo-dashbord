import enum
import hashlib
import re
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

from app.parser.utils import load_workbook_any


class HierarchyLevel(str, enum.Enum):
    BUYER = "buyer"
    CONTRACT = "contract"
    DOCUMENT = "document"
    TOTAL = "total"
    UNKNOWN = "unknown"


_INN_RE = re.compile(r"^\d{10}(\d{2})?$")
_PERIOD_RE = re.compile(
    r"за\s+(\w+)\s+(\d{4})\s*г?\.\s*-\s*(\w+)\s+(\d{4})\s*г?\.",
    re.IGNORECASE,
)
_DOC_NUM_DATE_RE = re.compile(r"№\s*(.+?)\s+от\s+(\d{2}\.\d{2}\.\d{4})")
_CONTRACT_NUM_DATE_RE = re.compile(r"№\s*(.+?)\s+от\s+(\d{2}\.\d{2}\.\d{4})")

_MONTHS = {
    "январь": 1, "февраль": 2, "март": 3, "апрель": 4,
    "май": 5, "июнь": 6, "июль": 7, "август": 8,
    "сентябрь": 9, "октябрь": 10, "ноябрь": 11, "декабрь": 12,
}


_CONTRACT_PREFIXES = (
    "Договор", "Основной договор", "ДОГОВОР",
    "Соглашение", "Допсоглашение", "ДС ", "Дополнительное соглашение",
    "Счет на оплату", "Счёт на оплату", "Счет-оферта", "Счёт-оферта",
    "СЧ ", "бн от", "бн  от",
)
_CONTRACT_LIKE_RE = re.compile(r"^(No?\d{3,}[-/]|\d{3,}[-/])")


def detect_level(name: str, inn_cell: str) -> HierarchyLevel:
    name = name.strip()
    inn_cell = str(inn_cell).strip() if inn_cell else ""

    if name.startswith("Итого"):
        return HierarchyLevel.TOTAL

    if inn_cell and _INN_RE.match(inn_cell):
        return HierarchyLevel.BUYER

    if name.startswith(_CONTRACT_PREFIXES) or _CONTRACT_LIKE_RE.match(name):
        return HierarchyLevel.CONTRACT

    if name.startswith(("Реализация", "Поступление")):
        return HierarchyLevel.DOCUMENT

    return HierarchyLevel.UNKNOWN


def _to_decimal(value) -> Decimal | None:
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _parse_date(date_str: str) -> date | None:
    try:
        parts = date_str.strip().split(".")
        return date(int(parts[2]), int(parts[1]), int(parts[0]))
    except (ValueError, IndexError):
        return None


def _detect_doc_type(name: str) -> str:
    if name.startswith("Реализация"):
        return "sale"
    if name.startswith("Поступление"):
        return "payment"
    return "unknown"


def _parse_period(text: str) -> tuple[date | None, date | None]:
    match = _PERIOD_RE.search(text)
    if not match:
        return None, None
    month_start = _MONTHS.get(match.group(1).lower())
    year_start = int(match.group(2))
    month_end = _MONTHS.get(match.group(3).lower())
    year_end = int(match.group(4))
    if not month_start or not month_end:
        return None, None
    start = date(year_start, month_start, 1)
    # Last day approximation: use first of next month - 1 day
    if month_end == 12:
        end = date(year_end, 12, 31)
    else:
        end = date(year_end, month_end + 1, 1).replace(day=1)
        from datetime import timedelta
        end = end - timedelta(days=1)
    return start, end


@dataclass
class ParsedDocument:
    raw_name: str
    doc_type: str
    doc_number: str | None = None
    doc_date: date | None = None
    amount: Decimal | None = None
    debt_start: Decimal | None = None
    advance_start: Decimal | None = None
    sold: Decimal | None = None
    paid: Decimal | None = None
    prepay_in: Decimal | None = None
    prepay_used: Decimal | None = None
    debt_end: Decimal | None = None
    advance_end: Decimal | None = None


@dataclass
class ParsedContract:
    raw_name: str
    contract_number: str | None = None
    contract_date: date | None = None
    debt_start: Decimal | None = None
    advance_start: Decimal | None = None
    sold: Decimal | None = None
    paid: Decimal | None = None
    prepay_in: Decimal | None = None
    prepay_used: Decimal | None = None
    debt_end: Decimal | None = None
    advance_end: Decimal | None = None
    documents: list[ParsedDocument] = field(default_factory=list)


@dataclass
class ParsedBuyer:
    name: str
    inn: str
    debt_start: Decimal | None = None
    advance_start: Decimal | None = None
    sold: Decimal | None = None
    paid: Decimal | None = None
    prepay_in: Decimal | None = None
    prepay_used: Decimal | None = None
    debt_end: Decimal | None = None
    advance_end: Decimal | None = None
    contracts: list[ParsedContract] = field(default_factory=list)


@dataclass
class ParseResult:
    filename: str
    file_hash: str
    period_start: date | None = None
    period_end: date | None = None
    total_rows: int = 0
    buyers_count: int = 0
    contracts_count: int = 0
    documents_count: int = 0
    buyers: list[ParsedBuyer] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def parse_debt_report(file_path: str | Path) -> ParseResult:
    file_path = Path(file_path)

    # Compute file hash
    file_hash = hashlib.sha256(file_path.read_bytes()).hexdigest()

    wb = load_workbook_any(file_path)
    ws = wb.active

    result = ParseResult(
        filename=file_path.name,
        file_hash=file_hash,
    )

    # Parse period from row 2
    period_cell = ws.cell(row=2, column=1).value
    if period_cell:
        result.period_start, result.period_end = _parse_period(str(period_cell))

    current_buyer: ParsedBuyer | None = None
    current_contract: ParsedContract | None = None

    # Data starts at row 9
    for row_idx in range(9, ws.max_row + 1):
        name_val = ws.cell(row=row_idx, column=1).value
        inn_val = ws.cell(row=row_idx, column=2).value

        if name_val is None:
            continue

        name_str = str(name_val).strip()
        inn_str = str(inn_val).strip() if inn_val is not None else ""

        level = detect_level(name_str, inn_str)

        if level == HierarchyLevel.TOTAL:
            break

        # Read column values: C=debt_start, D=advance_start, E=sold, F=paid,
        # G=prepay_in, H=prepay_used, I=debt_end, J=advance_end
        cols = {
            "debt_start": _to_decimal(ws.cell(row=row_idx, column=3).value),
            "advance_start": _to_decimal(ws.cell(row=row_idx, column=4).value),
            "sold": _to_decimal(ws.cell(row=row_idx, column=5).value),
            "paid": _to_decimal(ws.cell(row=row_idx, column=6).value),
            "prepay_in": _to_decimal(ws.cell(row=row_idx, column=7).value),
            "prepay_used": _to_decimal(ws.cell(row=row_idx, column=8).value),
            "debt_end": _to_decimal(ws.cell(row=row_idx, column=9).value),
            "advance_end": _to_decimal(ws.cell(row=row_idx, column=10).value),
        }

        result.total_rows += 1

        if level == HierarchyLevel.BUYER:
            current_buyer = ParsedBuyer(name=name_str, inn=inn_str, **cols)
            result.buyers.append(current_buyer)
            current_contract = None

        elif level == HierarchyLevel.CONTRACT:
            if current_buyer is None:
                result.errors.append(f"Row {row_idx}: contract without buyer")
                continue
            # Parse contract number and date
            num_match = _CONTRACT_NUM_DATE_RE.search(name_str)
            contract_number = num_match.group(1) if num_match else None
            contract_date = _parse_date(num_match.group(2)) if num_match else None
            current_contract = ParsedContract(
                raw_name=name_str,
                contract_number=contract_number,
                contract_date=contract_date,
                **cols,
            )
            current_buyer.contracts.append(current_contract)

        elif level == HierarchyLevel.DOCUMENT:
            if current_contract is None:
                result.errors.append(f"Row {row_idx}: document without contract")
                continue
            doc_type = _detect_doc_type(name_str)
            num_match = _DOC_NUM_DATE_RE.search(name_str)
            doc_number = num_match.group(1) if num_match else None
            doc_date = _parse_date(num_match.group(2)) if num_match else None
            # Amount: sold for sale, paid for payment
            amount = cols["sold"] if doc_type == "sale" else cols["paid"]
            doc = ParsedDocument(
                raw_name=name_str,
                doc_type=doc_type,
                doc_number=doc_number,
                doc_date=doc_date,
                amount=amount,
                **cols,
            )
            current_contract.documents.append(doc)
        else:
            result.errors.append(f"Row {row_idx}: unknown level for '{name_str}'")

    result.buyers_count = len(result.buyers)
    result.contracts_count = sum(len(b.contracts) for b in result.buyers)
    result.documents_count = sum(
        len(c.documents) for b in result.buyers for c in b.contracts
    )

    wb.close()
    return result
