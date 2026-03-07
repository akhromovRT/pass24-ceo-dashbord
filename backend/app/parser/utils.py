from pathlib import Path

import openpyxl


def load_workbook_any(file_path: str | Path) -> openpyxl.Workbook:
    """Load .xls or .xlsx file and return an openpyxl Workbook."""
    file_path = Path(file_path)
    if file_path.suffix.lower() == ".xls":
        import xlrd

        xls_wb = xlrd.open_workbook(str(file_path))
        wb = openpyxl.Workbook()
        for i, sheet_name in enumerate(xls_wb.sheet_names()):
            xls_ws = xls_wb.sheet_by_name(sheet_name)
            ws = wb.active if i == 0 else wb.create_sheet()
            ws.title = sheet_name
            for row in range(xls_ws.nrows):
                for col in range(xls_ws.ncols):
                    ws.cell(row=row + 1, column=col + 1, value=xls_ws.cell(row, col).value)
        return wb
    return openpyxl.load_workbook(file_path)
