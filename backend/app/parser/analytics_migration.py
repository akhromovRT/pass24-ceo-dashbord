"""Parser for the 'План vs Факт' sheet of the analytics spreadsheet.

Populates organizations with metadata (object, type, cloud_url, etc.),
creates contracts, and generates monthly snapshots (2025 actuals + 2026 plan).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import openpyxl

# --- Column mapping (1-indexed) ---
COL_INN = 1  # A
COL_NAME = 2  # B
COL_CONTRACT = 3  # C
COL_DATE = 4  # D
COL_AP = 5  # E
COL_PERIOD = 6  # F
COL_OBJECT = 7  # G
COL_OBJ_TYPE = 8  # H
COL_STATUS = 9  # I
COL_CLOUD = 10  # J
COL_SYSNUM = 11  # K
COL_EQUIPMENT = 12  # L
COL_ADDRESS = 13  # M
COL_CITY = 14  # N

# 2025 months: pairs of (plan_col, fact_col) starting at P(16)
# P/Q=Jan, R/S=Feb, T/U=Mar, V/W=Apr, X/Y=May, Z/AA=Jun,
# AB/AC=Jul, AD/AE=Aug, AF/AG=Sep, AH/AI=Oct, AJ/AK=Nov, AL/AM=Dec
_2025_PLAN_FACT = [
    (16, 17),  # Jan
    (18, 19),  # Feb
    (20, 21),  # Mar
    (22, 23),  # Apr
    (24, 25),  # May
    (26, 27),  # Jun
    (28, 29),  # Jul
    (30, 31),  # Aug
    (32, 33),  # Sep
    (34, 35),  # Oct
    (36, 37),  # Nov
    (38, 39),  # Dec
]

# 2026 months: single column each (plan only), AQ(43)..BB(54)
_2026_PLAN_COLS = list(range(43, 55))  # Jan=43 .. Dec=54

DATA_START_ROW = 3

_CONTRACT_NUMBER_RE = re.compile(r"^(№\s*)?(.+?)(\s+от\s+.+)?$")


@dataclass
class AnalyticsRow:
    inn: str
    name: str
    contract_raw: str | None = None
    contract_date: date | None = None
    monthly_ap: Decimal | None = None
    period: str | None = None
    object_name: str | None = None
    object_type: str | None = None
    status: str | None = None
    cloud_url: str | None = None
    system_number: str | None = None
    equipment: str | None = None
    address: str | None = None
    city_region: str | None = None
    # month -> (plan, fact) for 2025; month -> (plan, None) for 2026
    monthly_data: dict[tuple[int, int], tuple[Decimal | None, Decimal | None]] = field(
        default_factory=dict
    )


def _to_decimal(val: object) -> Decimal | None:
    if val is None:
        return None
    try:
        d = Decimal(str(val))
        return d if d != 0 else None
    except Exception:
        return None


def _to_date(val: object) -> date | None:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _cell(ws: object, row: int, col: int) -> object:
    return ws.cell(row=row, column=col).value  # type: ignore[union-attr]


def parse_analytics(file_path: str | Path) -> list[AnalyticsRow]:
    """Parse the 'План vs Факт' sheet and return list of AnalyticsRow."""
    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows: list[AnalyticsRow] = []

    for r in range(DATA_START_ROW, ws.max_row + 1):
        inn_raw = _cell(ws, r, COL_INN)
        if not inn_raw:
            continue
        inn = str(inn_raw).strip()
        if not inn or not any(c.isdigit() for c in inn):
            continue

        name = str(_cell(ws, r, COL_NAME) or "").strip()
        if not name:
            continue

        row = AnalyticsRow(
            inn=inn,
            name=name,
            contract_raw=str(_cell(ws, r, COL_CONTRACT) or "").strip() or None,
            contract_date=_to_date(_cell(ws, r, COL_DATE)),
            monthly_ap=_to_decimal(_cell(ws, r, COL_AP)),
            period=str(_cell(ws, r, COL_PERIOD) or "").strip() or None,
            object_name=str(_cell(ws, r, COL_OBJECT) or "").strip() or None,
            object_type=str(_cell(ws, r, COL_OBJ_TYPE) or "").strip() or None,
            status=str(_cell(ws, r, COL_STATUS) or "").strip() or None,
            cloud_url=str(_cell(ws, r, COL_CLOUD) or "").strip() or None,
            system_number=str(_cell(ws, r, COL_SYSNUM) or "").strip() or None,
            equipment=str(_cell(ws, r, COL_EQUIPMENT) or "").strip() or None,
            address=str(_cell(ws, r, COL_ADDRESS) or "").strip() or None,
            city_region=str(_cell(ws, r, COL_CITY) or "").strip() or None,
        )

        # 2025 plan/fact pairs
        for month_idx, (plan_col, fact_col) in enumerate(_2025_PLAN_FACT, start=1):
            plan = _to_decimal(_cell(ws, r, plan_col))
            fact = _to_decimal(_cell(ws, r, fact_col))
            if plan is not None or fact is not None:
                row.monthly_data[(2025, month_idx)] = (plan, fact)

        # 2026 plan only
        for month_idx, col in enumerate(_2026_PLAN_COLS, start=1):
            plan = _to_decimal(_cell(ws, r, col))
            if plan is not None:
                row.monthly_data[(2026, month_idx)] = (plan, None)

        rows.append(row)

    wb.close()
    return rows
