import re
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

from app.parser.period_extraction import extract_periods
from app.parser.utils import load_workbook_any

_INVOICE_RE = re.compile(
    r"(?:СЧЕТ(?:У|А)?|СЧ\.?)\s*№\s*([\w\-/]+)",
    re.IGNORECASE,
)
_CONTRACT_RE = re.compile(
    r"ДОГОВОР[А-ЯЁа-яё\s]*№\s*([\w\-/]+)",
    re.IGNORECASE,
)
_PERIOD_SLASH_RE = re.compile(r"(?:ЗА\s+)?(\d{2})\s*/\s*(\d{4})", re.IGNORECASE)

_MONTHS_MAP = {
    "январ": 1, "феврал": 2, "март": 3, "апрел": 4,
    "май": 5, "мая": 5, "мае": 5, "июн": 6, "июл": 7,
    "август": 8, "сентябр": 9, "октябр": 10, "ноябр": 11, "декабр": 12,
}
_MONTH_PERIOD_RE = re.compile(
    r"(?:ЗА|В)\s+(" + "|".join(_MONTHS_MAP.keys()) + r")[А-ЯЁа-яё]*\s+(\d{4})",
    re.IGNORECASE,
)

_TARIFF_RE = re.compile(r"\b(PROF|ПРОФ|STANDART|СТАНДАРТ)\b", re.IGNORECASE)
_TARIFF_MAP = {
    "PROF": "PROF", "ПРОФ": "PROF",
    "STANDART": "STANDART", "СТАНДАРТ": "STANDART",
}


@dataclass
class PaymentInfo:
    invoice_number: str | None = None
    contract_number: str | None = None
    period_month: int | None = None
    period_year: int | None = None
    tariff: str | None = None
    periods: list = field(default_factory=list)  # list[(year, month)]
    coverage_months: int | None = None
    payment_kind: str = "subscription"


@dataclass
class ParsedPayment:
    date: date
    doc_number: str
    amount: Decimal
    counterparty: str
    inn: str
    description: str
    doc_type: str | None = None
    payment_info: PaymentInfo = field(default_factory=PaymentInfo)


@dataclass
class BankStatementResult:
    filename: str
    account_number: str | None = None
    period: str | None = None
    owner: str | None = None
    payments: list[ParsedPayment] = field(default_factory=list)


def extract_payment_info(description: str) -> PaymentInfo:
    info = PaymentInfo()

    # Invoice number
    m = _INVOICE_RE.search(description)
    if m:
        info.invoice_number = m.group(1).rstrip(".")

    # Contract number
    m = _CONTRACT_RE.search(description)
    if m:
        info.contract_number = m.group(1).rstrip(".")

    # Period: try slash format first (03/2026)
    m = _PERIOD_SLASH_RE.search(description)
    if m:
        info.period_month = int(m.group(1))
        info.period_year = int(m.group(2))
    else:
        # Try month name format
        m = _MONTH_PERIOD_RE.search(description)
        if m:
            month_stem = m.group(1).lower()
            for stem, num in _MONTHS_MAP.items():
                if month_stem.startswith(stem):
                    info.period_month = num
                    break
            info.period_year = int(m.group(2))

    # Tariff
    m = _TARIFF_RE.search(description)
    if m:
        info.tariff = _TARIFF_MAP.get(m.group(1).upper())

    return info


def _parse_date_cell(value) -> date | None:
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            parts = value.strip().split(".")
            return date(int(parts[2]), int(parts[1]), int(parts[0]))
        except (ValueError, IndexError):
            return None
    return None


def parse_bank_statement(file_path: str | Path) -> BankStatementResult:
    file_path = Path(file_path)
    wb = load_workbook_any(file_path)
    ws = wb.active

    result = BankStatementResult(filename=file_path.name)

    # Parse header (rows 1-5)
    result.account_number = str(ws.cell(row=1, column=2).value or "").strip()
    result.period = str(ws.cell(row=2, column=2).value or "").strip()
    result.owner = str(ws.cell(row=3, column=2).value or "").strip()

    # Data starts at row 10
    for row_idx in range(10, ws.max_row + 1):
        date_val = ws.cell(row=row_idx, column=1).value
        if date_val is None:
            continue

        parsed_date = _parse_date_cell(date_val)
        if parsed_date is None:
            continue

        credit = ws.cell(row=row_idx, column=4).value
        if credit is None or credit == 0:
            continue  # Skip debit-only rows

        try:
            amount = Decimal(str(credit))
        except (InvalidOperation, ValueError):
            continue

        counterparty = str(ws.cell(row=row_idx, column=5).value or "").strip()
        inn = str(ws.cell(row=row_idx, column=6).value or "").strip()
        description = str(ws.cell(row=row_idx, column=11).value or "").strip()
        doc_number = str(ws.cell(row=row_idx, column=2).value or "").strip()
        doc_type = str(ws.cell(row=row_idx, column=13).value or "").strip() or None

        info = extract_payment_info(description)
        ep = extract_periods(description, parsed_date)
        info.periods = ep.periods
        info.coverage_months = ep.coverage_months
        info.payment_kind = ep.payment_kind
        if ep.periods:
            info.period_year, info.period_month = ep.periods[0]

        payment = ParsedPayment(
            date=parsed_date,
            doc_number=doc_number,
            amount=amount,
            counterparty=counterparty,
            inn=inn,
            description=description,
            doc_type=doc_type,
            payment_info=info,
        )
        result.payments.append(payment)

    wb.close()
    return result
