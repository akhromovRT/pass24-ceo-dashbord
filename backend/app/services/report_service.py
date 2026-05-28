"""Сборка управленческих отчётов по долгу и собираемости.

Три пресета:
- `debtors` — реестр должников: кого взыскивать, с приоритетом взыскания.
- `discipline` — дисциплина платежей: кто соскальзывает в долг.
- `payments` — реестр платежей за период (Софья 2026-05-28): дата платежа,
  контрагент, ИНН, сумма, назначение. Фильтры по диапазону дат и контрагентам.

Метрики собираемости берутся из AR-леджера (monthly_charges + payment_allocations),
а не из валовых сумм платежей — так же, как в dashboard.py."""
import uuid
from datetime import date
from io import BytesIO

from pydantic import BaseModel, ConfigDict
from sqlmodel import Session, func, select

from app.models import (
    Contract,
    DocType,
    Document,
    MonthlyCharge,
    Organization,
    OrgStatus,
    PaymentAllocation,
    User,
)
from app.services.aging import aging_index
from app.services.composition_service import (
    build_composition_report,
    columns_for_composition,
)

# --- критерии ---------------------------------------------------------------


class ReportCriteria(BaseModel):
    """Набор критериев отчёта. Сохраняется в ReportTemplate.criteria как есть."""

    model_config = ConfigDict(extra="ignore")

    period_from: str | None = None  # "YYYY-MM"
    period_to: str | None = None    # "YYYY-MM"
    statuses: list[str] = []
    manager_id: str | None = None
    aging_buckets: list[str] = []
    contract_types: list[str] = []
    city: str | None = None
    min_debt: float = 0.0
    include_excluded: bool = False
    columns: list[str] = []
    sort_by: str | None = None
    sort_dir: str = "desc"
    # только для пресета composition; остальные пресеты игнорируют через extra="ignore"
    metric: str | None = None
    period: str | None = None
    # P3.0.4 (Софья, 2026-05-28): точный диапазон дат и фильтр по контрагентам.
    # Используется в пресете payments; в остальных — игнорируется через extra="ignore".
    date_from: date | None = None
    date_to: date | None = None
    organization_ids: list[uuid.UUID] = []


# --- каталоги колонок -------------------------------------------------------

DEBTORS_COLUMNS: list[tuple[str, str]] = [
    ("name", "Клиент"),
    ("inn", "ИНН"),
    ("manager", "Менеджер"),
    ("monthly_ap", "АП/мес"),
    ("total_debt", "Долг"),
    ("months_overdue", "Месяцев просрочки"),
    ("aging_bucket", "Корзина"),
    ("last_payment_date", "Последний платёж"),
    ("collectability", "Собираемость, %"),
    ("payment_score", "Payment score"),
    ("priority", "Приоритет взыскания"),
    ("city", "Город"),
    ("status", "Статус"),
    ("churn_month", "Месяц отключения"),
]

DISCIPLINE_COLUMNS: list[tuple[str, str]] = [
    ("name", "Клиент"),
    ("inn", "ИНН"),
    ("manager", "Менеджер"),
    ("monthly_ap", "АП/мес"),
    ("collectability", "Собираемость, %"),
    ("on_time", "On-time, %"),
    ("tendency", "Тенденция"),
    ("trend", "Тренд"),
    ("unpaid_streak", "Неоплачено подряд, мес"),
    ("payment_score", "Payment score"),
    ("total_debt", "Долг"),
    ("status", "Статус"),
    ("churn_month", "Месяц отключения"),
]

PAYMENTS_COLUMNS: list[tuple[str, str]] = [
    ("doc_date", "Дата платежа"),
    ("org_name", "Контрагент"),
    ("inn", "ИНН"),
    ("amount", "Сумма, ₽"),
    ("raw_name", "Назначение"),
    ("allocated_periods", "Зачтено за"),
    ("manager", "Менеджер"),
]

REPORTS: dict[str, dict] = {
    "debtors": {"title": "Реестр должников", "columns": DEBTORS_COLUMNS},
    "discipline": {"title": "Дисциплина платежей", "columns": DISCIPLINE_COLUMNS},
    "composition": {"title": "Состав показателя", "columns": []},
    "payments": {"title": "Платежи", "columns": PAYMENTS_COLUMNS},
}

_RU_MONTHS = [
    "январь", "февраль", "март", "апрель", "май", "июнь",
    "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь",
]


def _fmt_churn_month(d) -> str:
    if d is None:
        return ""
    return f"{_RU_MONTHS[d.month - 1]} {d.year}"


STATUS_LABELS = {
    OrgStatus.ACTIVE: "Активен",
    OrgStatus.CHURNED: "Отток",
    OrgStatus.SUSPENDED: "Приостановлен",
    OrgStatus.PROSPECT: "Потенциальный",
    OrgStatus.TRANSIT: "Транзит",
}


# --- утилиты ----------------------------------------------------------------


def _f(x) -> float:
    return 0.0 if x is None else float(x)


def _parse_month(value: str | None) -> tuple[int, int] | None:
    """'2026-03' -> (2026, 3); мусор -> None."""
    if not value:
        return None
    try:
        y, m = value.split("-")[:2]
        yi, mi = int(y), int(m)
        if 1 <= mi <= 12:
            return (yi, mi)
    except (ValueError, AttributeError):
        pass
    return None


def _filter_orgs(session: Session, c: ReportCriteria) -> list[Organization]:
    """Применяет общие критерии (исключённые, менеджер, город, статус, тип
    договора). Базовые предикаты пресета (долг / подписка) — в build-функциях."""
    query = select(Organization)
    if not c.include_excluded:
        query = query.where(Organization.excluded_from_analytics == False)  # noqa: E712
    if c.manager_id:
        query = query.where(Organization.manager_id == c.manager_id)
    if c.city:
        query = query.where(
            func.lower(Organization.city_region).like(f"%{c.city.lower()}%")
        )
    statuses = [OrgStatus(s) for s in c.statuses if s in OrgStatus._value2member_map_]
    if statuses:
        query = query.where(Organization.status.in_(statuses))  # type: ignore[union-attr]
    if c.contract_types:
        org_ids = session.exec(
            select(Contract.organization_id)
            .where(Contract.contract_type.in_(c.contract_types))  # type: ignore[union-attr]
            .distinct()
        ).all()
        query = query.where(Organization.id.in_(org_ids))  # type: ignore[union-attr]
    return list(session.exec(query).all())


def _load_ledger(session: Session, org_ids: list) -> dict:
    """{org_id: [charge, ...]}; charge = {year, month, amount, allocs}.
    allocs — список (amount, pay_year, pay_month) с датой платежа."""
    if not org_ids:
        return {}
    charge_rows = session.exec(
        select(MonthlyCharge.id, MonthlyCharge.organization_id,
               MonthlyCharge.year, MonthlyCharge.month, MonthlyCharge.amount)
        .where(MonthlyCharge.organization_id.in_(org_ids))  # type: ignore[union-attr]
    ).all()
    charges: dict = {}
    by_id: dict = {}
    for cid, oid, y, m, amount in charge_rows:
        rec = {"year": int(y), "month": int(m), "amount": _f(amount), "allocs": []}
        charges.setdefault(oid, []).append(rec)
        by_id[cid] = rec

    alloc_rows = session.exec(
        select(PaymentAllocation.monthly_charge_id,
               PaymentAllocation.allocated_amount, Document.doc_date)
        .join(Document, Document.id == PaymentAllocation.payment_document_id)
        .join(MonthlyCharge, MonthlyCharge.id == PaymentAllocation.monthly_charge_id)
        .where(MonthlyCharge.organization_id.in_(org_ids))  # type: ignore[union-attr]
    ).all()
    for cid, amount, doc_date in alloc_rows:
        rec = by_id.get(cid)
        if rec is None:
            continue
        py = doc_date.year if doc_date else None
        pm = doc_date.month if doc_date else None
        rec["allocs"].append((_f(amount), py, pm))

    for recs in charges.values():
        recs.sort(key=lambda r: (r["year"], r["month"]))
    return charges


def _last_payments(session: Session, org_ids: list) -> dict:
    if not org_ids:
        return {}
    rows = session.exec(
        select(Document.organization_id, func.max(Document.doc_date))
        .where(Document.doc_type == DocType.PAYMENT,
               Document.organization_id.in_(org_ids))  # type: ignore[union-attr]
        .group_by(Document.organization_id)
    ).all()
    return {oid: d for oid, d in rows if d is not None}


def _managers(session: Session) -> dict:
    return {u.id: u.name for u in session.exec(select(User)).all()}


def _in_period(key: tuple[int, int], pf, pt, cur) -> bool:
    if pf and key < pf:
        return False
    if key > (pt or cur):
        return False
    return True


def _collectability(charges: list, pf, pt, cur) -> float | None:
    """Собираемость периода: Σ разнесённых платежей ÷ Σ начислений, %."""
    accrued = collected = 0.0
    for c in charges:
        if not _in_period((c["year"], c["month"]), pf, pt, cur):
            continue
        accrued += c["amount"]
        collected += sum(a[0] for a in c["allocs"])
    if accrued <= 0:
        return None
    return round(collected / accrued * 100, 1)


def _sort_rows(rows: list[dict], sort_by: str | None, sort_dir: str,
               default_key: str) -> list[dict]:
    key = sort_by or default_key
    if rows and key not in rows[0]:
        key = default_key

    def val(v):
        return v.lower() if isinstance(v, str) else v

    present = [r for r in rows if r.get(key) is not None]
    nulls = [r for r in rows if r.get(key) is None]
    present.sort(key=lambda r: val(r[key]), reverse=(sort_dir != "asc"))
    return present + nulls  # None всегда в конце


# --- пресет 1: реестр должников ---------------------------------------------


def build_debtors_report(session: Session, c: ReportCriteria) -> list[dict]:
    today = date.today()
    cur = (today.year, today.month)
    pf, pt = _parse_month(c.period_from), _parse_month(c.period_to)

    orgs = [o for o in _filter_orgs(session, c)
            if _f(o.total_debt) > max(c.min_debt, 0.0)]
    org_ids = [o.id for o in orgs]
    ledger = _load_ledger(session, org_ids)
    last_pay = _last_payments(session, org_ids)
    managers = _managers(session)
    aging = aging_index(session)  # возраст долга — по календарным месяцам

    rows: list[dict] = []
    for o in orgs:
        debt = _f(o.total_debt)
        monthly = _f(o.monthly_ap)
        bucket, age_m = aging.get(o.id, ("90+", 3))
        if c.aging_buckets and bucket not in c.aging_buckets:
            continue
        score = o.payment_score if o.payment_score is not None else 50
        priority = debt * (1 + min(age_m, 12) * 0.15) * (0.3 + 0.7 * score / 100)
        lp = last_pay.get(o.id)
        rows.append({
            "name": o.name_display or o.name_1c,
            "inn": o.inn,
            "manager": managers.get(o.manager_id, "—"),
            "monthly_ap": round(monthly, 2) if monthly else None,
            "total_debt": round(debt, 2),
            "months_overdue": age_m,
            "aging_bucket": bucket,
            "last_payment_date": lp.isoformat() if lp else None,
            "collectability": _collectability(ledger.get(o.id, []), pf, pt, cur),
            "payment_score": o.payment_score,
            "priority": round(priority, 2),
            "city": o.city_region or "—",
            "status": STATUS_LABELS.get(o.status, str(o.status)),
            "churn_month": _fmt_churn_month(o.churn_month),
        })
    return _sort_rows(rows, c.sort_by, c.sort_dir, "priority")


# --- пресет 2: дисциплина платежей ------------------------------------------


def _tendency(charges: list, pf, pt, cur) -> str:
    """Преобладающая тенденция оплаты по месяцу платежа vs месяцу начисления."""
    advance = current = arrears = 0.0
    for c in charges:
        ck = (c["year"], c["month"])
        if not _in_period(ck, pf, pt, cur):
            continue
        for amt, py, pm in c["allocs"]:
            if py is None:
                continue
            pk = (py, pm)
            if ck > pk:
                advance += amt
            elif ck < pk:
                arrears += amt
            else:
                current += amt
    if advance == current == arrears == 0:
        return "—"
    best = max((advance, "аванс"), (current, "в срок"), (arrears, "с просрочкой"),
               key=lambda t: t[0])
    return best[1]


def _trend(charges: list, pf, pt, cur) -> str:
    """Тренд собираемости: последние 3 месяца периода против 3 предыдущих."""
    months = sorted(
        (c for c in charges
         if _in_period((c["year"], c["month"]), pf, pt, cur)),
        key=lambda c: (c["year"], c["month"]),
    )
    if len(months) < 4:
        return "→"
    recent, older = months[-3:], months[-6:-3]
    if not older:
        return "→"

    def ratio(group: list) -> float | None:
        acc = sum(c["amount"] for c in group)
        col = sum(a[0] for c in group for a in c["allocs"])
        return col / acc * 100 if acc > 0 else None

    r, o = ratio(recent), ratio(older)
    if r is None or o is None:
        return "→"
    if r - o > 5:
        return "↑"
    if r - o < -5:
        return "↓"
    return "→"


def _unpaid_streak(charges: list, cur) -> int:
    """Сколько последних (по текущий месяц) начислений подряд не закрыты."""
    past = sorted(
        (c for c in charges if (c["year"], c["month"]) <= cur),
        key=lambda c: (c["year"], c["month"]), reverse=True,
    )
    streak = 0
    for c in past:
        outstanding = c["amount"] - sum(a[0] for a in c["allocs"])
        if outstanding > 0.01:
            streak += 1
        else:
            break
    return streak


def _on_time_pct(charges: list, pf, pt, cur) -> float | None:
    """Доля начислений периода, закрытых платежами не позже месяца начисления."""
    total = on_time = 0
    for c in charges:
        ck = (c["year"], c["month"])
        if not _in_period(ck, pf, pt, cur):
            continue
        total += 1
        paid_on_time = sum(
            amt for amt, py, pm in c["allocs"]
            if py is not None and (py, pm) <= ck
        )
        if paid_on_time >= c["amount"] - 0.01:
            on_time += 1
    if total == 0:
        return None
    return round(on_time / total * 100, 1)


def build_discipline_report(session: Session, c: ReportCriteria) -> list[dict]:
    today = date.today()
    cur = (today.year, today.month)
    pf, pt = _parse_month(c.period_from), _parse_month(c.period_to)

    # дисциплина — про подписчиков; по умолчанию активные, статусы переопределяемы
    orgs = [o for o in _filter_orgs(session, c) if _f(o.monthly_ap) > 0]
    if not c.statuses:
        orgs = [o for o in orgs if o.status == OrgStatus.ACTIVE]
    org_ids = [o.id for o in orgs]
    ledger = _load_ledger(session, org_ids)
    managers = _managers(session)

    rows: list[dict] = []
    for o in orgs:
        charges = ledger.get(o.id, [])
        rows.append({
            "name": o.name_display or o.name_1c,
            "inn": o.inn,
            "manager": managers.get(o.manager_id, "—"),
            "monthly_ap": round(_f(o.monthly_ap), 2),
            "collectability": _collectability(charges, pf, pt, cur),
            "on_time": _on_time_pct(charges, pf, pt, cur),
            "tendency": _tendency(charges, pf, pt, cur),
            "trend": _trend(charges, pf, pt, cur),
            "unpaid_streak": _unpaid_streak(charges, cur),
            "payment_score": o.payment_score,
            "total_debt": round(_f(o.total_debt), 2),
            "status": STATUS_LABELS.get(o.status, str(o.status)),
            "churn_month": _fmt_churn_month(o.churn_month),
        })
    # худшие (низкая собираемость) — сверху
    return _sort_rows(rows, c.sort_by, c.sort_dir, "collectability")


# --- диспетчер и экспорт ----------------------------------------------------

def build_payments_report(session: Session, c: ReportCriteria) -> list[dict]:
    """Реестр платежей за период с фильтрами по контрагентам и диапазону дат.

    Каждая строка — один PAYMENT-документ. Колонка `allocated_periods`
    показывает за какие месяцы зачтён платёж (по AR-леджеру), чтобы
    Софье было видно «дата прихода → за какие периоды засчитан»."""
    query = (
        select(Document, Organization)
        .join(Organization, Organization.id == Document.organization_id)
        .where(Document.doc_type == DocType.PAYMENT)
    )
    if not c.include_excluded:
        query = query.where(Organization.excluded_from_analytics == False)  # noqa: E712
    if c.date_from:
        query = query.where(Document.doc_date >= c.date_from)  # type: ignore[union-attr]
    if c.date_to:
        query = query.where(Document.doc_date <= c.date_to)  # type: ignore[union-attr]
    if c.organization_ids:
        query = query.where(Document.organization_id.in_(c.organization_ids))  # type: ignore[union-attr]
    if c.statuses:
        statuses = [OrgStatus(s) for s in c.statuses if s in {st.value for st in OrgStatus}]
        if statuses:
            query = query.where(Organization.status.in_(statuses))  # type: ignore[union-attr]
    query = query.order_by(Document.doc_date.desc())  # type: ignore[union-attr]

    pairs = list(session.exec(query).all())
    if not pairs:
        return []

    # Подтягиваем аллокации одним запросом — для колонки «Зачтено за».
    payment_ids = [d.id for d, _ in pairs]
    alloc_rows = session.exec(
        select(PaymentAllocation.payment_document_id,
               MonthlyCharge.year, MonthlyCharge.month,
               PaymentAllocation.allocated_amount)
        .join(MonthlyCharge,
              MonthlyCharge.id == PaymentAllocation.monthly_charge_id,
              isouter=True)
        .where(PaymentAllocation.payment_document_id.in_(payment_ids))  # type: ignore[union-attr]
    ).all()
    alloc_by_pay: dict[uuid.UUID, list[tuple[int, int, float]]] = {}
    for pay_id, y, m, amt in alloc_rows:
        if y is None or m is None:
            continue
        alloc_by_pay.setdefault(pay_id, []).append((y, m, float(amt)))

    managers = _managers(session)

    def _fmt_periods(items: list[tuple[int, int, float]]) -> str:
        # «04.2026 (12 000 ₽) + 05.2026 (3 000 ₽)»
        items.sort()
        return " + ".join(
            f"{m:02d}.{y} ({amt:,.0f} ₽)".replace(",", " ")
            for y, m, amt in items
        )

    out: list[dict] = []
    for doc, org in pairs:
        periods = alloc_by_pay.get(doc.id, [])
        out.append({
            "doc_date": doc.doc_date.isoformat() if doc.doc_date else None,
            "org_name": org.name_display or org.name_1c,
            "inn": org.inn,
            "amount": _f(doc.amount),
            "raw_name": doc.raw_name or "",
            "allocated_periods": _fmt_periods(periods) if periods else "не разнесено",
            "manager": managers.get(org.manager_id, "") if org.manager_id else "",
        })
    return _sort_rows(out, c.sort_by, c.sort_dir, default_key="doc_date")


_BUILDERS = {
    "debtors": build_debtors_report,
    "discipline": build_discipline_report,
    "composition": build_composition_report,
    "payments": build_payments_report,
}


def build_report(report_type: str, session: Session, c: ReportCriteria) -> list[dict]:
    builder = _BUILDERS.get(report_type)
    if builder is None:
        raise ValueError(f"unknown report type: {report_type}")
    return builder(session, c)


def columns_for(report_type: str, c: ReportCriteria) -> list[tuple[str, str]]:
    if report_type == "composition":
        return columns_for_composition(c)
    catalog = REPORTS[report_type]["columns"]
    if not c.columns:
        return catalog
    chosen = set(c.columns)
    selected = [pair for pair in catalog if pair[0] in chosen]
    return selected or catalog


def report_to_xlsx(report_type: str, columns: list[tuple[str, str]],
                    rows: list[dict]) -> bytes:
    """Собирает .xlsx: жирная цветная шапка, заморозка строки заголовков."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = REPORTS[report_type]["title"][:31]

    ws.append([header for _, header in columns])
    head_fill = PatternFill("solid", fgColor="6366F1")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([row.get(key) for key, _ in columns])

    for idx, (_, header) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(14, len(header) + 2)
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
