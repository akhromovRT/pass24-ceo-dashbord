"""Тесты /import/preview и /import/commit."""

import hashlib
import io
import json
from datetime import UTC, datetime

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlmodel import Session, select

from app.api.v1.auth import get_current_user
from app.core.database import get_session
from app.main import app
from app.models import Document, ImportRun, ImportStatus, Organization, User, UserRole


@pytest.fixture
def client(db_session: Session):
    def override_session():
        yield db_session

    def override_user():
        return User(
            name="T", email="t@t.ru", hashed_password="x", role=UserRole.ADMIN, is_active=True
        )

    app.dependency_overrides[get_session] = override_session
    app.dependency_overrides[get_current_user] = override_user
    yield TestClient(app)
    app.dependency_overrides.clear()


def _make_bank_xlsx(rows: list[dict]) -> bytes:
    """Минимальный xlsx в формате банк-выписки."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1, "Выписка по счёту")
    ws.cell(1, 2, "40702")
    ws.cell(2, 1, "За период")
    ws.cell(2, 2, "01.05.2026 - 18.05.2026")
    ws.cell(3, 1, "Владелец счёта")
    ws.cell(3, 2, "ОНВИ Сервис")
    ws.cell(8, 1, "Дата")
    ws.cell(8, 2, "Номер")
    ws.cell(8, 4, "Кредит")
    ws.cell(8, 5, "Контрагент")
    ws.cell(8, 6, "ИНН")
    ws.cell(8, 11, "Назначение")
    for i, r in enumerate(rows, start=10):
        ws.cell(i, 1, r["date"])
        ws.cell(i, 2, r["doc_number"])
        ws.cell(i, 4, r["amount"])
        ws.cell(i, 5, r["counterparty"])
        ws.cell(i, 6, r["inn"])
        ws.cell(i, 11, r.get("description", ""))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_file(content: bytes, name: str = "test.xlsx"):
    return (
        "file",
        (name, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    )


# Платёж БЕЗ периода в назначении
_NO_PERIOD_ROW = {
    "date": "18.05.2026",
    "doc_number": "542",
    "amount": 12387.38,
    "counterparty": 'ООО "Агростройкомплекс"',
    "inn": "5024039775",
    "description": "Доступ на месяц 2026г. к системе PASS24.online",
}
# Платёж С периодом в назначении
_WITH_PERIOD_ROW = {
    "date": "15.05.2026",
    "doc_number": "692",
    "amount": 12387.38,
    "counterparty": "ТСН ОРДЖОНИКИДЗЕ 1",
    "inn": "9725054546",
    "description": "ОПЛАТА ЗА ДОСТУП К СИСТЕМЕ PASS24.ONLINE STANDART МАЙ 2026 ГОДА",
}


class TestPreview:
    def test_preview_returns_payments_without_period(self, client):
        content = _make_bank_xlsx([_NO_PERIOD_ROW])
        resp = client.post(
            "/api/v1/import/preview?source_type=bank",
            files=[_xlsx_file(content)],
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["summary"]["total_payments"] == 1
        assert data["summary"]["without_period"] == 1
        assert len(data["payments"]) == 1
        p = data["payments"][0]
        assert p["index"] == 0
        assert p["detected_period"] is None
        assert p["inn"] == "5024039775"

    def test_preview_skips_payments_with_period(self, client):
        content = _make_bank_xlsx([_NO_PERIOD_ROW, _WITH_PERIOD_ROW])
        resp = client.post(
            "/api/v1/import/preview?source_type=bank",
            files=[_xlsx_file(content)],
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["summary"]["total_payments"] == 2
        assert data["summary"]["without_period"] == 1
        assert len(data["payments"]) == 1
        assert data["payments"][0]["index"] == 0

    def test_preview_returns_file_hash(self, client):
        content = _make_bank_xlsx([_NO_PERIOD_ROW])
        resp = client.post(
            "/api/v1/import/preview?source_type=bank",
            files=[_xlsx_file(content)],
        )
        data = resp.json()
        assert "file_hash" in data
        assert len(data["file_hash"]) == 64  # sha256 hex

    def test_preview_does_not_write_to_db(self, client, db_session: Session):
        content = _make_bank_xlsx([_NO_PERIOD_ROW])
        client.post(
            "/api/v1/import/preview?source_type=bank",
            files=[_xlsx_file(content)],
        )
        assert db_session.exec(select(ImportRun)).first() is None
        assert db_session.exec(select(Organization)).first() is None

    def test_preview_409_on_already_imported(self, client, db_session: Session):
        content = _make_bank_xlsx([_NO_PERIOD_ROW])
        file_hash = hashlib.sha256(content).hexdigest()
        run = ImportRun(
            filename="test.xlsx",
            file_hash=file_hash,
            status=ImportStatus.COMPLETED,
            completed_at=datetime.now(UTC),
        )
        db_session.add(run)
        db_session.commit()

        resp = client.post(
            "/api/v1/import/preview?source_type=bank",
            files=[_xlsx_file(content)],
        )
        assert resp.status_code == 409


class TestCommit:
    def test_commit_writes_to_db(self, client, db_session: Session):
        content = _make_bank_xlsx([_NO_PERIOD_ROW])
        file_hash = hashlib.sha256(content).hexdigest()

        resp = client.post(
            "/api/v1/import/commit?source_type=bank",
            files=[_xlsx_file(content)],
            data={"file_hash": file_hash, "period_overrides_json": "{}"},
        )
        assert resp.status_code == 200
        assert db_session.exec(select(ImportRun)).first() is not None

    def test_commit_applies_period_override(self, client, db_session: Session):
        content = _make_bank_xlsx([_NO_PERIOD_ROW])
        file_hash = hashlib.sha256(content).hexdigest()
        overrides = json.dumps({"0": {"year": 2026, "month": 5}})

        resp = client.post(
            "/api/v1/import/commit?source_type=bank",
            files=[_xlsx_file(content)],
            data={"file_hash": file_hash, "period_overrides_json": overrides},
        )
        assert resp.status_code == 200

        docs = db_session.exec(select(Document)).all()
        assert len(docs) == 1
        assert docs[0].period_manual is True
        assert docs[0].period_year == 2026
        assert docs[0].period_month == 5

    def test_commit_400_on_hash_mismatch(self, client):
        content = _make_bank_xlsx([_NO_PERIOD_ROW])

        resp = client.post(
            "/api/v1/import/commit?source_type=bank",
            files=[_xlsx_file(content)],
            data={"file_hash": "wrong_hash_aaaaaa", "period_overrides_json": "{}"},
        )
        assert resp.status_code == 400

    def test_commit_payments_source_type(self, client, db_session: Session):
        """POST /commit?source_type=payments с минимальным валидным payments-xlsx."""
        wb = openpyxl.Workbook()
        ws = wb.active
        # Строка 1 — заголовок (12 колонок)
        ws.append(
            [
                "Есть файлы",
                "Номер",
                "Дата",
                "Поступление",
                "Списание",
                "Контрагент",
                "ИНН",
                "Договор",
                "Вх.номер",
                "Вх.дата",
                "Назначение платежа",
                "Вид операции",
            ]
        )
        # Строка 2 — один платёж (с периодом в назначении)
        ws.append(
            [
                "Нет",
                "00БП-1",
                "15.05.2026",
                12100.0,
                None,
                "ООО Ромашка",
                "7700000001",
                "Договор №1",
                "5",
                "15.05.2026",
                "Оплата за доступ к PASS24 за май 2026",
                "Оплата от покупателя",
            ]
        )
        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()
        file_hash = hashlib.sha256(content).hexdigest()

        resp = client.post(
            "/api/v1/import/commit?source_type=payments",
            files=[_xlsx_file(content, "payments.xlsx")],
            data={"file_hash": file_hash, "period_overrides_json": "{}"},
        )
        assert resp.status_code == 200
        assert db_session.exec(select(ImportRun)).first() is not None

    def test_commit_400_on_malformed_overrides_json(self, client):
        content = _make_bank_xlsx([_NO_PERIOD_ROW])
        file_hash = hashlib.sha256(content).hexdigest()

        resp = client.post(
            "/api/v1/import/commit?source_type=bank",
            files=[_xlsx_file(content)],
            data={"file_hash": file_hash, "period_overrides_json": "not-json"},
        )
        assert resp.status_code == 400

    def test_commit_422_on_out_of_range_period(self, client):
        content = _make_bank_xlsx([_NO_PERIOD_ROW])
        file_hash = hashlib.sha256(content).hexdigest()
        overrides = json.dumps({"0": {"year": 2026, "month": 13}})

        resp = client.post(
            "/api/v1/import/commit?source_type=bank",
            files=[_xlsx_file(content)],
            data={"file_hash": file_hash, "period_overrides_json": overrides},
        )
        assert resp.status_code == 422
