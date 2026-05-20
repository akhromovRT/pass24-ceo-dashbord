import io
from datetime import date
from decimal import Decimal

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlmodel import Session

from app.api.v1.auth import get_current_user
from app.core.database import get_session
from app.main import app
from app.models import (
    AllocationBasis,
    ChargeSource,
    Contract,
    ContractType,
    DocType,
    Document,
    MonthlyCharge,
    Organization,
    OrgStatus,
    PaymentAllocation,
    User,
    UserRole,
)


@pytest.fixture
def client(db_session: Session):
    def override_get_session():
        yield db_session

    def override_get_current_user():
        return User(name="T", email="t@t.ru", hashed_password="x",
                    role=UserRole.ADMIN, is_active=True)

    app.dependency_overrides[get_session] = override_get_session
    app.dependency_overrides[get_current_user] = override_get_current_user
    yield TestClient(app)
    app.dependency_overrides.clear()


# --- хелперы посева ---------------------------------------------------------


def _org(session: Session, inn: str, name: str = "Org", **kw) -> Organization:
    org = Organization(inn=inn, name_1c=name, **kw)
    session.add(org)
    session.commit()
    session.refresh(org)
    return org


def _charge(session: Session, org: Organization, year: int, month: int,
            amount: float) -> MonthlyCharge:
    charge = MonthlyCharge(organization_id=org.id, year=year, month=month,
                           amount=Decimal(str(amount)),
                           source=ChargeSource.SYNTHETIC_TARIFF)
    session.add(charge)
    session.commit()
    session.refresh(charge)
    return charge


def _pay(session: Session, org: Organization, charge: MonthlyCharge,
         amount: float, pay_date: date) -> None:
    contract = Contract(organization_id=org.id,
                        contract_type=ContractType.SUBSCRIPTION)
    session.add(contract)
    session.commit()
    session.refresh(contract)
    doc = Document(contract_id=contract.id, organization_id=org.id,
                   doc_type=DocType.PAYMENT, doc_date=pay_date,
                   amount=Decimal(str(amount)))
    session.add(doc)
    session.commit()
    session.refresh(doc)
    alloc = PaymentAllocation(payment_document_id=doc.id,
                              monthly_charge_id=charge.id,
                              allocated_amount=Decimal(str(amount)),
                              basis=AllocationBasis.EXPLICIT_PERIOD)
    session.add(alloc)
    session.commit()


# --- пресет «реестр должников» ----------------------------------------------


def test_debtors_preview_returns_rows_and_columns(client, db_session: Session):
    _org(db_session, "7700000001", name="Должник",
         status=OrgStatus.ACTIVE, monthly_ap=Decimal("10000"),
         total_debt=Decimal("45000"), payment_score=60)
    resp = client.post("/api/v1/reports/debtors/preview", json={})
    assert resp.status_code == 200
    body = resp.json()
    assert body["total"] == 1
    row = body["rows"][0]
    assert row["total_debt"] == 45000.0
    assert row["aging_bucket"] == "90+"
    assert row["priority"] > 0
    keys = {c["key"] for c in body["columns"]}
    assert {"priority", "aging_bucket", "collectability"} <= keys


def test_debtors_min_debt_filter(client, db_session: Session):
    _org(db_session, "7700000010", name="Мелкий", total_debt=Decimal("5000"))
    _org(db_session, "7700000011", name="Крупный", total_debt=Decimal("80000"))
    resp = client.post("/api/v1/reports/debtors/preview",
                        json={"min_debt": 10000})
    assert resp.status_code == 200
    rows = resp.json()["rows"]
    assert len(rows) == 1
    assert rows[0]["name"] == "Крупный"


def test_debtors_sort_ascending(client, db_session: Session):
    _org(db_session, "7700000020", name="A", total_debt=Decimal("30000"))
    _org(db_session, "7700000021", name="B", total_debt=Decimal("90000"))
    resp = client.post("/api/v1/reports/debtors/preview",
                        json={"sort_by": "total_debt", "sort_dir": "asc"})
    rows = resp.json()["rows"]
    assert [r["total_debt"] for r in rows] == [30000.0, 90000.0]


def test_debtors_column_selection(client, db_session: Session):
    _org(db_session, "7700000030", name="X", total_debt=Decimal("12000"))
    resp = client.post("/api/v1/reports/debtors/preview",
                        json={"columns": ["name", "total_debt"]})
    cols = [c["key"] for c in resp.json()["columns"]]
    assert cols == ["name", "total_debt"]


# --- пресет «дисциплина платежей» -------------------------------------------


def test_discipline_preview_computes_collectability(client, db_session: Session):
    today = date.today()
    org = _org(db_session, "7700000040", name="Подписчик",
               status=OrgStatus.ACTIVE, monthly_ap=Decimal("10000"))
    charge = _charge(db_session, org, today.year, today.month, 10000)
    _pay(db_session, org, charge, 10000, today.replace(day=1))
    resp = client.post("/api/v1/reports/discipline/preview", json={})
    assert resp.status_code == 200
    rows = resp.json()["rows"]
    assert len(rows) == 1
    row = rows[0]
    assert row["collectability"] == 100.0
    assert row["tendency"] == "в срок"
    assert row["unpaid_streak"] == 0


def test_discipline_excludes_non_subscription(client, db_session: Session):
    # без monthly_ap клиент не попадает в отчёт о дисциплине подписчиков
    _org(db_session, "7700000050", name="Без подписки",
         status=OrgStatus.ACTIVE, total_debt=Decimal("9000"))
    resp = client.post("/api/v1/reports/discipline/preview", json={})
    assert resp.status_code == 200
    assert resp.json()["total"] == 0


# --- экспорт и валидация ----------------------------------------------------


def test_export_returns_valid_xlsx(client, db_session: Session):
    _org(db_session, "7700000060", name="Экспорт",
         monthly_ap=Decimal("10000"), total_debt=Decimal("33000"))
    resp = client.post("/api/v1/reports/debtors/export", json={})
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    workbook = load_workbook(io.BytesIO(resp.content))
    sheet = workbook.active
    assert sheet["A1"].value == "Клиент"
    assert sheet.max_row == 2  # шапка + одна строка данных


def test_unknown_report_type_returns_404(client):
    resp = client.post("/api/v1/reports/unknown/preview", json={})
    assert resp.status_code == 404


# --- шаблоны критериев ------------------------------------------------------


def test_template_crud(client, db_session: Session):
    created = client.post("/api/v1/reports/templates", json={
        "name": "Крупные должники",
        "report_type": "debtors",
        "criteria": {"min_debt": 50000, "sort_by": "priority"},
    })
    assert created.status_code == 201
    template_id = created.json()["id"]

    listed = client.get("/api/v1/reports/templates")
    assert listed.status_code == 200
    assert len(listed.json()) == 1
    assert listed.json()[0]["criteria"]["min_debt"] == 50000

    deleted = client.delete(f"/api/v1/reports/templates/{template_id}")
    assert deleted.status_code == 200
    assert client.get("/api/v1/reports/templates").json() == []


def test_template_rejects_unknown_type(client):
    resp = client.post("/api/v1/reports/templates", json={
        "name": "bad", "report_type": "nope", "criteria": {},
    })
    assert resp.status_code == 404


# --- пресет «состав показателя» ---------------------------------------------


def test_composition_preset_registered(client):
    """Composition зарегистрирован; preview возвращает 200 даже на пустых данных."""
    resp = client.post("/api/v1/reports/composition/preview",
                        json={"metric": "mrr_plan"})
    assert resp.status_code == 200
    body = resp.json()
    assert body["rows"] == []
    assert body["total"] == 0


def test_composition_unknown_metric_returns_400(client):
    resp = client.post("/api/v1/reports/composition/preview",
                        json={"metric": "bogus"})
    assert resp.status_code == 400


def test_composition_missing_metric_returns_400(client):
    """Запрос без metric → 400 с понятным сообщением."""
    resp = client.post("/api/v1/reports/composition/preview", json={})
    assert resp.status_code == 400
    assert "metric is required" in resp.json()["detail"].lower()
